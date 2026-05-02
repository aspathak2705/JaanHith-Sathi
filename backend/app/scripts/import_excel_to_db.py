import argparse
from pathlib import Path
import re
from typing import Any, Dict, Optional

from openpyxl import load_workbook

from app.db.init_db import init_db
from app.db.location_models import Booth
from app.db.session import SessionLocal

BACKEND_DIR = Path(__file__).resolve().parents[2]
DEFAULT_DATA_DIR = BACKEND_DIR / "app" / "data" / "polling Station Data"
HEADER_MARKERS = {
    "polling station name",
    "booth name",
    "मतदान केंद्र",
    "मतदान केंद्राचा तपशील",
}


def _value(row: Dict[str, Any], *names: str) -> Optional[Any]:
    for name in names:
        value = row.get(name)
        if value not in (None, ""):
            return value
    return None


def _text(row: Dict[str, Any], *names: str) -> str:
    value = _value(row, *names)
    if value is None:
        return ""
    return str(value).strip()


def _float(row: Dict[str, Any], *names: str) -> Optional[float]:
    value = _value(row, *names)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _area_from_booth_name(booth_name: str) -> str:
    parts = booth_name.split("-", 1)
    if len(parts) == 2:
        return parts[1].strip()
    return booth_name.strip()


def _room_from_detail(detail: str) -> str:
    match = re.search(r"(room\s*(?:no|number)?\.?\s*\S+|खोली\s*क्र\.?\s*[\w०-९]+)", detail, re.IGNORECASE)
    return match.group(1).strip() if match else ""


def _building_from_detail(detail: str) -> str:
    if not detail:
        return ""
    return re.split(r"\(,\)|\s+-\s+", detail, maxsplit=1)[0].strip(" ,-")


def _rows_from_xlsx(file_path: Path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active

    header_row = None
    header = None
    for idx, values in enumerate(sheet.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
        normalized = [
            str(column).strip().lower() if column is not None else ""
            for column in values
        ]
        if any(column in HEADER_MARKERS for column in normalized):
            header_row = idx
            header = normalized
            break

    if not header_row or not header:
        workbook.close()
        return

    try:
        for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            row = dict(zip(header, values))
            if any(value not in (None, "") for value in row.values()):
                yield row
    finally:
        workbook.close()


def import_excel_to_db(data_dir: Path = DEFAULT_DATA_DIR, clear_existing: bool = False) -> int:
    if not data_dir.exists():
        raise FileNotFoundError(f"Polling station data directory not found: {data_dir}")

    init_db()
    db = SessionLocal()
    imported = 0

    try:
        if clear_existing:
            db.query(Booth).delete()

        for district_path in sorted(data_dir.iterdir()):
            if not district_path.is_dir():
                continue

            for file_path in sorted(district_path.glob("*.xlsx")):
                city = file_path.stem

                for row in _rows_from_xlsx(file_path):
                    booth_name = _text(
                        row,
                        "polling station name",
                        "booth name",
                        "booth_name",
                        "मतदान केंद्र",
                    )
                    detail = _text(
                        row,
                        "polling station details",
                        "polling station detail",
                        "मतदान केंद्राचा तपशील",
                    )
                    building = _text(row, "building name", "building")
                    area = _text(row, "area", "locality")
                    room = _text(row, "room no", "room", "room number")

                    if not building:
                        building = _building_from_detail(detail)
                    if not area:
                        area = _area_from_booth_name(booth_name)
                    if not room:
                        room = _room_from_detail(detail)

                    db.add(
                        Booth(
                            district=district_path.name,
                            city=city,
                            booth_name=booth_name,
                            building=building,
                            area=area,
                            room=room,
                            latitude=_float(row, "latitude", "lat"),
                            longitude=_float(row, "longitude", "lng", "lon", "long"),
                        )
                    )
                    imported += 1

        db.commit()
        return imported
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import polling station Excel data into SQLite.")
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--clear", action="store_true", help="Clear existing booth rows before import.")
    args = parser.parse_args()

    count = import_excel_to_db(args.data_dir, clear_existing=args.clear)
    print(f"Imported {count} polling booths into SQLite.")
